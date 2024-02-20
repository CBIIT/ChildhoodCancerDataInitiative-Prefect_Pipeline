import pandas as pd
import os
import openpyxl
from datetime import date
import warnings
from openpyxl.utils.dataframe import dataframe_to_rows
from src.utils import get_logger, get_date, get_time
from prefect import flow
import requests
import re


@flow(
    name="CCDI_to_CDS_conversion",
    flow_run_name="CCDI_to_CDS_conversion_" + f"{get_time()}",
)
def CCDI_to_CDS(manifest_path: str) -> tuple:
    # pull in args as variables
    file_path = manifest_path

    # get logger
    logger = get_logger(loggername="CCDI_to_CDS_submission", log_level="info")
    logger.info("The CCDI to CDS conversion has begun")

    # Download CDS model template from GitHub repo
    cds_template_url = (
        "https://api.github.com/repos/CBIIT/cds-model/contents/metadata-manifest/"
    )
    cds_template_content = requests.get(cds_template_url).json()
    if re.search(
        "v(0|[1-9]\d*)\.(0|[1-9]\d*)\.(0|[1-9]\d*)\.xlsx$",
        cds_template_content[0]["name"],
    ):
        cds_template_dl_url = cds_template_content[0]["download_url"]
        template_dl_res = requests.get(cds_template_dl_url)
        template_file_name = cds_template_content[0]["name"]
        template_file = open(template_file_name, "wb")
        template_file.write(template_dl_res.content)
        logger.info(f"Downloaded CDS template manifest {template_file_name}")
        template_path = template_file_name
        template_version = template_file_name.strip("\.xlsx").split("_")[-1]
        logger.info(f"CDS template manifest version is {template_version}")
    else:
        logger.info(
            "Couldn't download the CDS template manifest. Please visit https://github.com/CBIIT/cds-model/tree/main/metadata-manifest"
        )

    ##############
    #
    # File name rework
    #
    ##############

    # Determine file ext and abs path
    file_name = os.path.splitext(os.path.split(os.path.relpath(file_path))[1])[0]
    file_ext = os.path.splitext(file_path)[1]
    file_dir_path = os.path.split(os.path.abspath(file_path))[0]

    if file_dir_path == "":
        file_dir_path = "."

    # obtain the date
    def refresh_date():
        today = date.today()
        today = today.strftime("%Y%m%d")
        return today

    todays_date = refresh_date()

    # Output file name based on input file name and date/time stamped.
    output_file = file_name + "_CDS" + todays_date

    ##############
    #
    # Pull Dictionary Page to create node pulls
    #
    ##############

    def read_xlsx(file_path: str, sheet: str):
        # Read in excel file
        warnings.simplefilter(action="ignore", category=UserWarning)
        return pd.read_excel(file_path, sheet, dtype=str)

    # create workbook
    cds_model = pd.ExcelFile(template_path)

    # create CDS workbook
    cds_df = read_xlsx(cds_model, "Metadata")

    cds_df_dict = read_xlsx(cds_model, "Dictionary")

    cds_req_props = (
        cds_df_dict[cds_df_dict["Required"].notna()]["Field"].dropna().unique().tolist()
    )

    # close cds_model object
    cds_model.close()

    ##############
    #
    # Read in data
    #
    ##############

    # create workbook
    ccdi_data = pd.ExcelFile(file_path)

    # create dictionary for dfs
    ccdi_dfs = {}

    # read in dfs and apply to dictionary
    for sheet_name in ccdi_data.sheet_names:
        ccdi_dfs[sheet_name] = read_xlsx(ccdi_data, sheet_name)

    # ccdi nodes minus information nodes
    info_nodes = ["README and INSTRUCTIONS", "Dictionary", "Terms and Value Sets"]
    ccdi_nodes = [item for item in ccdi_data.sheet_names if item not in info_nodes]

    ## Go through each tab and remove completely empty tabs
    nodes_removed = []

    for node in ccdi_nodes:
        if node in ccdi_dfs:
            # see if the tab contain any data
            test_df = ccdi_dfs[node]
            test_df = test_df.drop("type", axis=1)
            test_df = test_df.dropna(how="all").dropna(how="all", axis=1)
            # if there is no data, drop the node/tab
            if test_df.empty:
                nodes_removed.append(node)
            else:
                pass
        else:
            nodes_removed.append(node)

    logger.info(f"{nodes_removed} tabs are empty")
    ccdi_dfs = {key: ccdi_dfs[key] for key in ccdi_dfs if key not in nodes_removed}    

    if "cell_line" not in nodes_removed or "pdx" not in nodes_removed:
        logger.warning(
            "This submission contains 'cell_line' or 'pdx' entries in the submission.\n\tTHIS MAY CAUSE ERRORS IN THE RUN OR RETURNED SUBMISSION FILE.\n\t\tDOUBLE CHECK THAT ALL DATA IS PRESENT IN THE OUTPUT."
        )

    # This was removed as the nodes required for CDS destroys paths that need to be walked to obtain the full data.
    ccdi_to_cds_nodes = [node for node in ccdi_nodes if node not in nodes_removed]
    print(f"{*ccdi_to_cds_nodes,}")

    ### MERGING OF ALL DATA
    # The variable names will be the initials of the node as they are added
    # This will show the addition order in hopes to keep the graph walking logic correct and consistent.
    # We are opting for this more structured hard coded approach as more fuzzy walking logics tend to fail or eat up memory.

    # function to make dropping of columns easy and customizable.
    def drop_type_id_others(data_frame, others_list=[]):
        if "type" in data_frame.columns:
            data_frame = data_frame.drop(["type"], axis=1)
        if "id" in data_frame.columns:
            data_frame = data_frame.drop(["id"], axis=1)
        if any(data_frame.columns.str.contains("\.id")):
            dot_id_cols = data_frame.columns[
                data_frame.columns.str.contains("\.id")
            ].tolist()
            for dot_id_col in dot_id_cols:
                data_frame = data_frame.drop([dot_id_col], axis=1)
        if others_list:
            for other in others_list:
                if other in data_frame.columns:
                    data_frame = data_frame.drop([other], axis=1)
        return data_frame

    # start with study
    df_all = drop_type_id_others(ccdi_dfs["study"])

    # make note of columns to remap, so that joins are easier
    col_remap = {
        "study.study_id": "study_id",
        "participant.participant_id": "participant_id",
        "sample.sample_id": "sample_id",
        "pdx.pdx_id": "pdx_id",
        "cell_line.cell_line_id": "cell_line_id",
    }

    # add study_admin
    if "study_admin" in ccdi_to_cds_nodes:
        df_node = drop_type_id_others(ccdi_dfs["study_admin"])
        df_node.rename(columns=col_remap, inplace=True)
        df_all = pd.merge(df_all, df_node, how="left", on="study_id")

    # add study_personnel
    if "study_personnel" in ccdi_to_cds_nodes:
        df_node = drop_type_id_others(ccdi_dfs["study_personnel"])
        df_node.rename(columns=col_remap, inplace=True)
        df_all = pd.merge(df_all, df_node, how="left", on="study_id")

    # pull out df for study
    df_study_level = df_all
    print("below is the df_all")
    print(df_all)

    # add participant
    if "participant" in ccdi_to_cds_nodes:
        df_node = drop_type_id_others(ccdi_dfs["participant"])
        df_node.rename(columns=col_remap, inplace=True)
        df_all = pd.merge(df_all, df_node, how="left", on="study_id")

    # add diagnosis
    if "diagnosis" in ccdi_to_cds_nodes:
        df_node = drop_type_id_others(ccdi_dfs["diagnosis"])
        df_node.rename(columns=col_remap, inplace=True)
        df_all = pd.merge(df_all, df_node, how="left", on="participant_id")

    # pull out df for diagnosis
    df_participant_level = df_all

    # ALL [node]_file nodes will need to be concatenated first so there are no conflicts on common column names:
    # file_name, file_type, dcf_indexd_guid, etc etc etc

    df_file = pd.DataFrame()

    if "radiology_file" in ccdi_to_cds_nodes:
        df_file = pd.concat([df_file, ccdi_dfs["radiology_file"]], ignore_index=True)

    if "sequencing_file" in ccdi_to_cds_nodes:
        df_file = pd.concat([df_file, ccdi_dfs["sequencing_file"]], ignore_index=True)

    if "methylation_array_file" in ccdi_to_cds_nodes:
        df_file = pd.concat(
            [df_file, ccdi_dfs["methylation_array_file"]], ignore_index=True
        )

    if "cytogenomic_file" in ccdi_to_cds_nodes:
        df_file = pd.concat([df_file, ccdi_dfs["cytogenomic_file"]], ignore_index=True)

    if "pathology_file" in ccdi_to_cds_nodes:
        df_file = pd.concat([df_file, ccdi_dfs["pathology_file"]], ignore_index=True)

    if "single_cell_sequencing_file" in ccdi_to_cds_nodes:
        df_file = pd.concat(
            [df_file, ccdi_dfs["single_cell_sequencing_file"]], ignore_index=True
        )

    if "clinical_measure_file" in ccdi_to_cds_nodes:
        df_file = pd.concat(
            [df_file, ccdi_dfs["clinical_measure_file"]], ignore_index=True
        )

    # rename the columns based on the col_remap dictionary made earlier
    df_file.rename(columns=col_remap, inplace=True)

    # drop off all the extra properties that are not required for transformation into a flattened data frame
    df_file = drop_type_id_others(df_file)

    # Check df_file to see if there are any rows
    # abort the script if no rows are left.
    if df_file.shape[0] == 0:
        logger.error(
            "No files were found in the submission template. Please add files or ignore the output from this step."
        )
        output_file_path = "(EMPTY)_"+ output_file + ".xlsx"
        logger_file_name = "CCDI_to_CDS_submission_" + get_date() + ".log"
        return (output_file_path, logger_file_name)
    else:
        pass

    # START WITH FILES INSTEAD AND WALK EACH LINE BACK?
    # Based on each connection possible for the files walk it back manually by starting with the most likely connections

    def join_node(parent_df, child_df, join_by):
        # get rid of book-keeping columns and rename linking columns to node level
        parent_df = drop_type_id_others(parent_df)
        parent_df.rename(columns=col_remap, inplace=True)
        child_df = drop_type_id_others(child_df)
        child_df.rename(columns=col_remap, inplace=True)
        # get rid of rows that are empty for the join by key
        child_df = child_df.dropna(subset=[join_by])
        # merge data frames
        df_joined = pd.merge(parent_df, child_df, how="left", on=join_by)
        return df_joined

    def join_file_node_cleaner(node_df):
        # clean up data frame if there are _x and _y columns
        # always giving preference to _x data frame which should be the parent data frame
        for col in node_df.columns.tolist():
            if col.endswith("_x"):
                col_x = col
                col_base = col[:-2]
                col_y = col_base + "_y"
                node_df[col_base] = node_df[col_x].combine_first(node_df[col_y])
                node_df.drop(columns=[col_x, col_y], inplace=True)
        print(node_df)
        # clean up the data frame, drop empty columns, rows that don't have files, reset index and remove duplicates.
        if "file_url_in_cds" in node_df.columns:
            print(node_df["file_url_in_cds"])
            node_df = node_df.dropna(subset=["file_url_in_cds"])
        print("dropna of empty file_url_in_cds")
        print(node_df)
        node_df = node_df.dropna(axis=1, how="all").reset_index(drop=True)
        print("dropna of all empty columns")
        print(node_df)
        #node_df = node_df.reset_index(drop=True)
        node_df = node_df.drop_duplicates()
        print("drop duplicates")
        print(node_df)
        return node_df

    # Do an empty check on the parent nodes before trying to join them
    # Since the final concatenation needs to have an object, each data frame will be made before the check to ensure it exists,
    # it will be removed from the final addition if it is an empty data frame.

    # file --> sample
    sample_file = pd.DataFrame()
    print(f"df_file columns: \n {*df_file.columns.tolist(),}")
    print(f"ccdi_dfs[sample]")
    print(ccdi_dfs["sample"])
    if "sample" in ccdi_to_cds_nodes:
        if "sample_id" in df_file.columns:
            sample_file = join_node(ccdi_dfs["sample"], df_file, "sample_id")
            print("printing sample_file after join_node")
            print(sample_file)
            sample_file = join_file_node_cleaner(sample_file)
            print("printing sample_file afer join_file_node_cleaner")
            print(sample_file)
    print("printing final sample_file df")
    print(sample_file)
    # file --> pdx
    pdx_file = pd.DataFrame()
    if "pdx" in ccdi_to_cds_nodes:
        if "pdx_id" in df_file.columns:
            pdx_file = join_node(ccdi_dfs["pdx"], df_file, "pdx_id")
            pdx_file = join_file_node_cleaner(pdx_file)

    # file --> cell_line
    cell_line_file = pd.DataFrame()
    if "cell_line" in ccdi_to_cds_nodes:
        if "cell_line_id" in df_file.columns:
            cell_line_file = join_node(ccdi_dfs["cell_line"], df_file, "cell_line_id")
            cell_line_file = join_file_node_cleaner(cell_line_file)

    # file --> participant
    participant_file = pd.DataFrame()
    if not df_participant_level.empty:
        if "participant_id" in df_file.columns:
            participant_file = join_node(
                df_participant_level, df_file, "participant_id"
            )
            participant_file = join_file_node_cleaner(participant_file)

    # file --> study
    study_file = pd.DataFrame()
    if not df_study_level.empty:
        if "study_id" in df_file.columns:
            study_file = join_node(df_study_level, df_file, "study_id")
            study_file = join_file_node_cleaner(study_file)

    # Then for the nodes that don't end on either sample, participant or study, walk again and check.
    # This is likely to be deprecated later as having all files go to sample is the preferred method.

    # pdx_file --> sample
    sample_pdx_file = pd.DataFrame()
    if not pdx_file.empty:
        if "sample_id" in pdx_file.columns:
            sample_pdx_file = join_node(ccdi_dfs["sample"], pdx_file, "sample_id")
            sample_pdx_file = join_file_node_cleaner(sample_pdx_file)

    # pdx_file --> study
    study_pdx_file = pd.DataFrame()
    if not pdx_file.empty:
        if "study_id" in pdx_file.columns:
            study_pdx_file = join_node(df_study_level, pdx_file, "study_id")
            study_pdx_file = join_file_node_cleaner(study_pdx_file)

    # cell_line_file --> sample
    sample_cell_line_file = pd.DataFrame()
    if not cell_line_file.empty:
        if "sample_id" in cell_line_file.columns:
            sample_cell_line_file = join_node(
                ccdi_dfs["sample"], cell_line_file, "sample_id"
            )
            sample_cell_line_file = join_file_node_cleaner(sample_cell_line_file)

    # cell_line_file --> participant
    participant_cell_line_file = pd.DataFrame()
    if not cell_line_file.empty:
        if "participant_id" in cell_line_file.columns:
            participant_cell_line_file = join_node(
                df_participant_level, cell_line_file, "participant_id"
            )
            participant_cell_line_file = join_file_node_cleaner(
                participant_cell_line_file
            )

    # cell_line_file --> study
    study_cell_line_file = pd.DataFrame()
    if not cell_line_file.empty:
        if "study_id" in cell_line_file.columns:
            study_cell_line_file = join_node(df_study_level, cell_line_file, "study_id")
            study_cell_line_file = join_file_node_cleaner(study_cell_line_file)

    # If not all samples can link to a participant, then handle those exceptions

    # sample_file --> pdx
    pdx_sample_file = pd.DataFrame()
    if not sample_file.empty:
        if "pdx_id" in sample_file.columns:
            pdx_sample_file = join_node(ccdi_dfs["pdx"], sample_file, "pdx_id")
            pdx_sample_file = join_file_node_cleaner(pdx_sample_file)

    # sample_file--> cell_line
    cell_line_sample_file = pd.DataFrame()
    if not sample_file.empty:
        if "cell_line_id" in sample_file.columns:
            cell_line_sample_file = join_node(
                ccdi_dfs["cell_line"], sample_file, "cell_line_id"
            )
            cell_line_sample_file = join_file_node_cleaner(cell_line_sample_file)

    # pdx_sample_file --> sample
    sample_pdx_sample_file = pd.DataFrame()
    # For entries that go from file-->sample-->[pdx/cell_line]-->sample
    # We have to remove the previous sample_ids in the file as they are confusing the join

    if not pdx_sample_file.empty:
        if "sample_id" in pdx_sample_file.columns:
            sample_pdx_sample_file = join_node(
                ccdi_dfs["sample"], pdx_sample_file, "sample_id"
            )
            sample_pdx_sample_file = join_file_node_cleaner(sample_pdx_sample_file)

    # pdx_sample_file --> study
    study_pdx_sample_file = pd.DataFrame()
    if not pdx_file.empty:
        if "study_id" in pdx_sample_file.columns:
            study_pdx_sample_file = join_node(
                df_study_level, pdx_sample_file, "study_id"
            )
            study_pdx_sample_file = join_file_node_cleaner(study_pdx_sample_file)

    # cell_line_sample_file --> sample
    sample_cell_line_sample_file = pd.DataFrame()
    if not cell_line_sample_file.empty:
        if "sample_id" in cell_line_sample_file.columns:
            sample_cell_line_sample_file = join_node(
                ccdi_dfs["sample"], cell_line_sample_file, "sample_id"
            )
            sample_cell_line_sample_file = join_file_node_cleaner(
                sample_cell_line_sample_file
            )

    # cell_line_sample_file --> participant
    participant_cell_line_sample_file = pd.DataFrame()
    if not cell_line_sample_file.empty:
        if "participant_id" in cell_line_sample_file.columns:
            participant_cell_line_sample_file = join_node(
                df_participant_level, cell_line_sample_file, "participant_id"
            )
            participant_cell_line_sample_file = join_file_node_cleaner(
                participant_cell_line_sample_file
            )

    # cell_line_sample_file --> study
    study_cell_line_sample_file = pd.DataFrame()
    if not cell_line_sample_file.empty:
        if "study_id" in cell_line_sample_file.columns:
            study_cell_line_sample_file = join_node(
                df_study_level, cell_line_sample_file, "study_id"
            )
            study_cell_line_sample_file = join_file_node_cleaner(
                study_cell_line_sample_file
            )

    # It is not likely to loop more than once as that would mean a pdx or cell_line was made from a sample of a pdx or cell_line.

    # Then handle the concatenation of the three main node end points:
    # Get everything up to sample, participant or study.
    # For all samples, we will join them with participant, which inherently has study information.
    # Then we will concatenate all the dfs into one large df, clean it up and we should get the same number of unique files.
    # Flattening files might cause duplicate rows, but this duplication is likely to leave when we pull out elements that are not usually multiple elements.

    ### INFO ###
    # List of Outputs from the different joins
    # # file --> sample
    # sample_file
    # # file --> pdx
    # pdx_file
    # # file --> cell_line
    # cell_line_file
    # # file --> participant
    # participant_file
    # # file --> study
    # study_file
    # # pdx_file --> sample
    # sample_pdx_file
    # # pdx_file --> study
    # study_pdx_file
    # # cell_line_file --> sample
    # sample_cell_line_file
    # # cell_line_file --> participant
    # participant_cell_line_file
    # # cell_line_file --> study
    # study_cell_line_file
    # # sample_file --> pdx
    # pdx_sample_file
    # # sample_file--> cell_line
    # cell_line_sample_file
    # # pdx_sample_file --> sample
    # sample_pdx_sample_file
    # # pdx_sample_file --> study
    # study_pdx_sample_file
    # # cell_line_sample_file --> sample
    # sample_cell_line_sample_file
    # # cell_line_sample_file --> participant
    # participant_cell_line_sample_file
    # # cell_line_sample_file --> study
    # study_cell_line_sample_file
    ### INFO ###

    ### INFO ###
    # List of outputs from the different joins sorted by end points

    # #PDX
    # # file --> pdx
    # pdx_file
    # # sample_file --> pdx
    # pdx_sample_file

    # #Cell_line
    # # file --> cell_line
    # cell_line_file
    # # sample_file--> cell_line
    # cell_line_sample_file

    # #Sample
    # # file --> sample
    # sample_file
    # # pdx_file --> sample
    # sample_pdx_file
    # # cell_line_file --> sample
    # sample_cell_line_file
    # # pdx_sample_file --> sample
    # sample_pdx_sample_file
    # # cell_line_sample_file --> sample
    # sample_cell_line_sample_file

    # #Participant
    # # file --> participant
    # participant_file
    # # cell_line_file --> participant
    # participant_cell_line_file
    # # cell_line_sample_file --> participant
    # participant_cell_line_sample_file

    # #Study
    # # file --> study
    # study_file
    # # pdx_file --> study
    # study_pdx_file
    # # cell_line_file --> study
    # study_cell_line_file
    # # pdx_sample_file --> study
    # study_pdx_sample_file
    # # cell_line_sample_file --> study
    # study_cell_line_sample_file
    ### INFO ###

    # The PDX and Cell_line outputs can be ignored as they must go to sample, participant or study.
    # The participant and study can be set aside as they should be able to be concatenated at the end.

    # Take the samples and attach to df_participant_level or df_study_level

    # Due to sample_type being a property that CCDI does not have and instead CCDI anatomic_site is used instead
    # we run into an issue where there are two anatomic_site properties that are being used, one in sample and the other in diagnosis.
    # In this context it makes more sense to use the samples.anatomic_site, which is counter to how the script handles most other conflicts.
    # The script treats the parent node's properties as the value to overwrite with, but in this one case we have to switch this to be more inline
    # with CDS, and hopefully prevent duplicate values.

    # Thus for the following node joins before the join_file_node_cleaner, part of that code will be used to do the reverse clean up ONLY for anatomic_site
    # and only when it is linking to the participant level.

    # sample_file --> participant
    participant_sample_file = pd.DataFrame()
    if not sample_file.empty:
        if "participant_id" in sample_file.columns:
            participant_sample_file = join_node(
                df_participant_level, sample_file, "participant_id"
            )

            # Handle the reverse situation of anatomic_site
            if "anatomic_site_x" in participant_sample_file.columns:
                col_x = "anatomic_site_x"
                col_base = col_x[:-2]
                col_y = col_base + "_y"
                participant_sample_file[col_base] = participant_sample_file[
                    col_y
                ].combine_first(participant_sample_file[col_x])
                participant_sample_file.drop(columns=[col_x, col_y], inplace=True)

            participant_sample_file = join_file_node_cleaner(participant_sample_file)

    # sample_pdx_file --> participant
    participant_sample_pdx_file = pd.DataFrame()
    if not sample_pdx_file.empty:
        if "participant_id" in sample_pdx_file.columns:
            participant_sample_pdx_file = join_node(
                df_participant_level, sample_pdx_file, "participant_id"
            )

            # Handle the reverse situation of anatomic_site
            if "anatomic_site_x" in participant_sample_pdx_file.columns:
                col_x = "anatomic_site_x"
                col_base = col_x[:-2]
                col_y = col_base + "_y"
                participant_sample_pdx_file[col_base] = participant_sample_pdx_file[
                    col_y
                ].combine_first(participant_sample_pdx_file[col_x])
                participant_sample_pdx_file.drop(columns=[col_x, col_y], inplace=True)

            participant_sample_pdx_file = join_file_node_cleaner(
                participant_sample_pdx_file
            )

    # sample_cell_line_file --> participant
    participant_sample_cell_line_file = pd.DataFrame()
    if not sample_cell_line_file.empty:
        if "participant_id" in sample_cell_line_file.columns:
            participant_sample_cell_line_file = join_node(
                df_participant_level, sample_cell_line_file, "participant_id"
            )

            # Handle the reverse situation of anatomic_site
            if "anatomic_site_x" in participant_sample_cell_line_file.columns:
                col_x = "anatomic_site_x"
                col_base = col_x[:-2]
                col_y = col_base + "_y"
                participant_sample_cell_line_file[
                    col_base
                ] = participant_sample_cell_line_file[col_y].combine_first(
                    participant_sample_cell_line_file[col_x]
                )
                participant_sample_cell_line_file.drop(
                    columns=[col_x, col_y], inplace=True
                )

            participant_sample_cell_line_file = join_file_node_cleaner(
                participant_sample_cell_line_file
            )

    # sample_pdx_sample_file --> participant
    participant_sample_pdx_sample_file = pd.DataFrame()
    if not sample_pdx_sample_file.empty:
        if "participant_id" in sample_pdx_sample_file.columns:
            participant_sample_pdx_sample_file = join_node(
                df_participant_level, sample_pdx_sample_file, "participant_id"
            )

            # Handle the reverse situation of anatomic_site
            if "anatomic_site_x" in participant_sample_pdx_sample_file.columns:
                col_x = "anatomic_site_x"
                col_base = col_x[:-2]
                col_y = col_base + "_y"
                participant_sample_pdx_sample_file[
                    col_base
                ] = participant_sample_pdx_sample_file[col_y].combine_first(
                    participant_sample_pdx_sample_file[col_x]
                )
                participant_sample_pdx_sample_file.drop(
                    columns=[col_x, col_y], inplace=True
                )

            participant_sample_pdx_sample_file = join_file_node_cleaner(
                participant_sample_pdx_sample_file
            )

    # sample_cell_line_sample_file --> participant
    participant_sample_cell_line_sample_file = pd.DataFrame()
    if not sample_cell_line_sample_file.empty:
        if "participant_id" in sample_cell_line_sample_file.columns:
            participant_sample_cell_line_sample_file = join_node(
                df_participant_level, sample_cell_line_sample_file, "participant_id"
            )

            # Handle the reverse situation of anatomic_site
            if "anatomic_site_x" in participant_sample_cell_line_sample_file.columns:
                col_x = "anatomic_site_x"
                col_base = col_x[:-2]
                col_y = col_base + "_y"
                participant_sample_cell_line_sample_file[
                    col_base
                ] = participant_sample_cell_line_sample_file[col_y].combine_first(
                    participant_sample_cell_line_sample_file[col_x]
                )
                participant_sample_cell_line_sample_file.drop(
                    columns=[col_x, col_y], inplace=True
                )

            participant_sample_cell_line_sample_file = join_file_node_cleaner(
                participant_sample_cell_line_sample_file
            )

    # sample_pdx_sample_file --> study
    study_sample_pdx_sample_file = pd.DataFrame()
    if not sample_pdx_sample_file.empty:
        if "study_id" in sample_pdx_sample_file.columns:
            study_sample_pdx_sample_file = join_node(
                df_study_level, sample_pdx_sample_file, "study_id"
            )
            study_sample_pdx_sample_file = join_file_node_cleaner(
                study_sample_pdx_sample_file
            )

    # List of all paths that can be derived from files that either end at study or participant (which has study information)
    all_paths = [
        participant_file,
        participant_cell_line_file,
        participant_cell_line_sample_file,
        participant_sample_file,
        participant_sample_pdx_file,
        participant_sample_cell_line_file,
        participant_sample_pdx_sample_file,
        participant_sample_cell_line_sample_file,
        study_file,
        study_pdx_file,
        study_cell_line_file,
        study_pdx_sample_file,
        study_cell_line_sample_file,
        study_sample_pdx_sample_file,
    ]

    df_join_all = pd.DataFrame()

    for node_path in all_paths:
        if not node_path.empty:
            df_join_all = pd.concat([df_join_all, node_path], axis=0, ignore_index=True)
            print(node_path)
        else:
            print(node_path)
            print("is empty")

    # To reduce complexity in the conversion, only lines where the personnel type is PI will be used in the CDS template end file. Otherwise use Co-PI or just pass if nothing else or too complex.
    if 'PI' in df_join_all['personnel_type'].unique().tolist():
        df_join_all=df_join_all[df_join_all['personnel_type']=='PI']
    elif 'Co-PI' in df_join_all['personnel_type'].unique().tolist():
        df_join_all=df_join_all[df_join_all['personnel_type']=='Co-PI']
    else:
        pass

    # Drop the current index as it is causing issues
    df_join_all = df_join_all.reset_index(drop=True)

    # To try and preserve synonym links this section will join the synonyms if the data exists.
    if 'synonym' in ccdi_to_cds_nodes:
        synonym_df = ccdi_dfs["synonym"]
        synonym_df.rename(columns=col_remap, inplace=True)

        if "participant_id" in synonym_df.columns:
            df_join_all = join_node(df_join_all, synonym_df, "participant_id")
            df_join_all = join_file_node_cleaner(df_join_all)
            # now we need to move participant_ids that are related to dbGaP over to the property `dbGaP_subject_id`
            # if statment because we are not asssured that links are correctly made
            if "repository_of_synonym_id" in df_join_all.columns:
                df_join_all["dbGaP_subject_id"] = df_join_all.loc[
                    df_join_all["repository_of_synonym_id"] == "dbGaP", "synonym_id"
                ]

        if "synonym_id" in df_join_all.columns:
            df_join_all = df_join_all.drop("synonym_id", axis=1)
        if "repository_of_synonym_id" in df_join_all.columns:
            df_join_all = df_join_all.drop("repository_of_synonym_id", axis=1)

        if "sample_id" in synonym_df.columns:
            print(f"synonym_df columns: {*synonym_df.columns.tolist(),}" )
            print(f"df_join_all columns: {*df_join_all.columns.tolist(),}")
            df_join_all = join_node(df_join_all, synonym_df, "sample_id")
            df_join_all = join_file_node_cleaner(df_join_all)
            # now we need to move sample_ids that are related to BioSample over to the property `dbGaP_subject_id`
            # if statment because we are not asssured that links are correctly made
            if "repository_of_synonym_id" in df_join_all.columns:
                df_join_all["biosample_accession"] = df_join_all.loc[
                    df_join_all["repository_of_synonym_id"] == "BioSample", "synonym_id"
                ]

    # Drop the current index as it is causing issues
    df_join_all = df_join_all.reset_index(drop=True)

    ###############
    #
    # CCDI to CDS required columns hard coding
    #
    ###############

    # From the df_join_all data frame, make either 1:1 mappings, rework mappings for different property names
    # or transform the data to match the new mapping setups.

    # for simple 1:1 mappings even if the property names are different, a simple function to handle it:
    def simple_add(cds_prop, ccdi_prop):
        if ccdi_prop in df_join_all:
            cds_df[cds_prop] = df_join_all[ccdi_prop]

        return

    # study and study modifiers
    simple_add("phs_accession", "phs_accession")
    simple_add("study_acronym", "study_acronym")
    simple_add("acl", "acl")
    simple_add("email", "email_address")
    simple_add("role_or_affiliation", "personnel_type")

    # NOT REQUIRED in CCDI, but REQUIRED IN CDS and can be derived via logic
    if "authz" in df_join_all.columns:
        cds_df["authz"] = df_join_all["authz"]
    else:
        authz = df_join_all["acl"].dropna().unique().tolist()[0]
        authz = "['/programs/" + authz[2:]
        cds_df["authz"] = authz

    # if there is one experimental value
    if (
        'experimental_strategy_and_data_subtype' in df_join_all.columns
    ):
        if (
            len(
                df_join_all["experimental_strategy_and_data_subtype"]
                .dropna()
                .unique()
                .tolist()
            )
            == 1
        ):
            cds_df["experimental_strategy_and_data_subtype"] = df_join_all[
                "experimental_strategy_and_data_subtype"
            ]
            # if there are multiple experimental values
        elif (
            len(
                df_join_all["experimental_strategy_and_data_subtype"]
                .dropna()
                .unique()
                .tolist()
            )
            > 1
        ):
            es_and_ds = ";".join(
                df_join_all["experimental_strategy_and_data_subtype"].unique().tolist()
            )
            cds_df["experimental_strategy_and_data_subtype"] = es_and_ds
            # if there are no experimental values
    else :
        cds_df["experimental_strategy_and_data_subtype"] = "Sequencing"

    # if there is one study data type value
    if (
        'study_data_types' in df_join_all.columns
    ):
        if len(df_join_all["study_data_types"].dropna().unique().tolist()) == 1:
            cds_df["study_data_types"] = df_join_all["study_data_types"]
            # if there are multiple study data type values
        elif len(df_join_all["study_data_types"].dropna().unique().tolist()) > 1:
            es_and_ds = ";".join(df_join_all["study_data_types"].unique().tolist())
            cds_df["study_data_types"] = es_and_ds
            # if there are no study data type values
    else:
        cds_df["study_data_types"] = "Genomics"

    # if there is a study_name
    if (
        'study_name' in df_join_all.columns
    ):
        cds_df["study_name"] = df_join_all["study_name"]
        # if there isn't a study_name
    else:
        cds_df["study_name"] = df_join_all["study_short_title"]

    # if there is a number_of_participants value
    if "number_of_participants" in df_join_all.columns:
        # if there is a number_of_participants
        if len(df_join_all["number_of_participants"].dropna().unique().tolist()) == 1:
            cds_df["number_of_participants"] = df_join_all["number_of_participants"]
            # if there isn't a number_of_participants
        if len(df_join_all["number_of_participants"].dropna().unique().tolist()) != 1:
            cds_df["number_of_participants"] = 1
    else:
        cds_df["number_of_participants"] = 1

    # if there is a number_of_samples value
    if "number_of_samples" in df_join_all.columns:
        # if there is a number_of_samples
        if len(df_join_all["number_of_samples"].dropna().unique().tolist()) == 1:
            cds_df["number_of_samples"] = df_join_all["number_of_samples"]
            # if there isn't a number_of_samples
        if len(df_join_all["number_of_samples"].dropna().unique().tolist()) != 1:
            cds_df["number_of_samples"] = 1
    else:
        cds_df["number_of_samples"] = 1

        # REQUIRED in CCDI, but has to be reworked

    # Has been reworked to eliminate all non-PI entities earlier in the script, thus this next line can be more simple.
    # personnel_names=df_join_all[df_join_all['personnel_type']=='PI']['personnel_name'].dropna().unique().tolist()
    personnel_names = df_join_all["personnel_name"].dropna().unique().tolist()

    for personnel_name in personnel_names:
        # clear previous entry
        title = None
        first = None
        middle = None
        last = None

        # create a true/false data frame to determine which rows get the name
        name_apply = (df_join_all["personnel_name"] == personnel_name).tolist()

        personnel_name = personnel_name.split(" ")
        prefix_delete = False
        # Need to keep prefix as title, just remove for determining name simplicity
        prefixes = [
            "Dr.",
            "Dr",
            "Mr.",
            "Mr",
            "Mrs.",
            "Mrs",
            "Ms.",
            "Ms",
            "Miss",
            "Sir",
            "Dame",
            "Lord",
            "Lady",
        ]
        first_name_part = personnel_name[0]

        if first_name_part in prefixes:
            prefix_delete = True
            title = first_name_part

        if prefix_delete:
            del personnel_name[0]

        if len(personnel_name) > 2:
            first = personnel_name[0]
            middle = personnel_name[1]
            last = " ".join(personnel_name[2:])
        elif len(personnel_name) == 2:
            first = personnel_name[0]
            last = personnel_name[1]
        elif len(personnel_name) == 1:
            last = personnel_name[0]

        # Apply names to each row that are applicable
        for x in range(0, len(name_apply)):
            if name_apply[x]:
                if title is not None:
                    cds_df.loc[x, "title"] = title
                cds_df.loc[x, "first_name"] = first
                cds_df.loc[x, "middle_name"] = middle
                cds_df.loc[x, "last_name"] = last

    # participant
    simple_add("participant_id", "participant_id")

    # diagnosis
    # Not a perfect match, some logic required depending on version of CCDI
    # Also need to deconstruct the setup from `code : term` ----> `term`
    if "diagnosis_icd_o" in df_join_all.columns:
        cds_df["primary_diagnosis"] = df_join_all["diagnosis_icd_o"]
    elif "diagnosis_classification" in df_join_all.columns:
        cds_df['primary_diagnosis']=df_join_all['diagnosis_classification']
        # further diagnosis handling for "see diagnosis_comment" copying
        if 'diagnosis_comment' in df_join_all.columns:
            comment_true = cds_df['primary_diagnosis'] == "see diagnosis_comment"
            cds_df.loc[comment_true, 'primary_diagnosis'] = df_join_all.loc[comment_true, 'diagnosis_comment']
    else:
        logger.error("No 'primary_diagnosis' was transfered")

    # sample
    simple_add("sample_id", "sample_id")
    # anatomic site is the closest approximation we can get, but it will be wrong most likely
    simple_add("sample_type", "anatomic_site")
    # this script will also put CCDI anatomic site in sample_anatomic_site for CDS as that is a better fit
    simple_add("sample_anatomic_site", "anatomic_site")

    # file
    simple_add("file_name", "file_name")
    simple_add("file_size", "file_size")
    simple_add("file_type", "file_type")
    simple_add("file_url_in_cds", "file_url_in_cds")
    simple_add("instrument_model", "instrument_model")
    simple_add("library_id", "library_id")
    simple_add("library_layout", "library_layout")
    simple_add("library_selection", "library_selection")
    simple_add("library_source", "library_source")
    simple_add("library_strategy", "library_strategy")
    simple_add("md5sum", "md5sum")
    simple_add("platform", "platform")
    simple_add("design_description", "design_description")

    # Not required in CCDI (further logic needed)

    # If it is there, it gets added, if not it will be changed to "Not Applicable" by later transformation
    simple_add("reference_genome_assembly", "reference_genome_assembly")

    # Not required, but "easy" data adds

    if "gender" in df_join_all:
        simple_add("gender", "gender")
    elif "sex_at_birth" in df_join_all:
        simple_add("gender", "sex_at_birth")

    simple_add("race", "race")
    simple_add("ethnicity", "ethnicity")
    simple_add("bases", "number_of_bp")
    simple_add("number_of_reads", "number_of_reads")
    simple_add("avg_read_length", "avg_read_length")
    simple_add("coverage", "coverage")
    simple_add("file_mapping_level", "file_mapping_level")
    simple_add("adult_or_childhood_study", "adult_or_childhood_study")
    simple_add("organism_species", "organism_species")
    simple_add("methylation_platform", "methylation_platform")
    simple_add("reporter_label", "reporter_label")
    simple_add("sample_description", "sample_description")
    simple_add("sample_tumor_status", "sample_tumor_status")
    simple_add("sequence_alignment_software", "sequence_alignment_software")
    simple_add("tumor_grade", "tumor_grade")
    simple_add("tumor_stage_clinical_t", "tumor_stage_clinical_t")
    simple_add("tumor_stage_clinical_n", "tumor_stage_clinical_n")
    simple_add("tumor_stage_clinical_m", "tumor_stage_clinical_m")
    simple_add("dbGaP_subject_id", "dbGaP_subject_id")
    simple_add("biosample_accession", "biosample_accession")
    simple_add("guid", "dcf_indexd_guid")

    # This property has been removed as it is not required by CDS and it is often a source for multiple lines for the same file
    # simple_add('age_at_diagnosis','age_at_diagnosis')

    # Remove any rows where there is not a file associated with the entry
    cds_df = cds_df.dropna(subset=["file_url_in_cds"])

    # Minor fix, if sample_id has no value, then sample_type/sample_anatomic_site should not have a value.
    cds_df.loc[cds_df["sample_id"].isnull(), "sample_type"] = None
    cds_df.loc[cds_df["sample_id"].isnull(), "sample_anatomic_site"] = None

    # Removed from output, see above
    # Minor fix, if age_at_diagnosis is unknown in CCDI (-999), remove that value as it is not required for CDS.
    # cds_df.loc[cds_df['age_at_diagnosis']=='-999', 'age_at_diagnosis'] = None

    # The not applicable transformation that takes any NAs in the data frame and applies "Not Applicable"
    # to the fields that are missing this required data.

    cds_df[cds_req_props] = cds_df[cds_req_props].fillna("Not Applicable")

    cds_df = cds_df.drop_duplicates()

    # Quick stats to check the conversion, make sure things are working as we think they should be.
    file_expected = len(
        df_file.loc[:, ["md5sum", "file_name", "file_url_in_cds"]].drop_duplicates()
    )
    file_returned = len(
        cds_df.loc[:, ["md5sum", "file_name", "file_url_in_cds"]].drop_duplicates()
    )

    logger.info(
        f"For the following conversion:\n\tUnique files expected: {file_expected}\n\tUnique files returned: {file_returned}"
    )

    if file_expected == file_returned:
        logger.info(
            "The conversion was likely a success, please double check your files"
        )
    elif file_expected != file_returned:
        logger.warning(
            "The conversion was likely NOT successful, as it did not return the same number of unique files"
        )

    ##############
    #
    # Write out
    #
    ##############

    logger.info("Writing out the CDS workbook file")

    template_workbook = openpyxl.load_workbook(template_path)

    ws = template_workbook["Metadata"]
    # remove any data that might be in the template
    ws.delete_rows(2, ws.max_row)

    # write the data
    for row in dataframe_to_rows(cds_df, index=False, header=False):
        ws.append(row)

    # add the GUID column to the template
    sheet = template_workbook["Metadata"]

    sheet.cell(row=1, column=len(cds_df.columns), value="guid")

    # save out template
    template_workbook.save(f"{file_dir_path}/{output_file}.xlsx")
    template_workbook.close()

    # output file path
    output_file_path = output_file + ".xlsx"

    logger.info(
        f"Process Complete. The output file can be found here: {file_dir_path}/{output_file}"
    )

    # logger file name
    logger_file_name = "CCDI_to_CDS_submission_" + get_date() + ".log"

    return (output_file_path, logger_file_name)
