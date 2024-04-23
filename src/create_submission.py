import yaml
from dataclasses import dataclass
from pathlib import Path
import numpy as np
import requests
import json
import pandas as pd
from openpyxl.utils import get_column_letter, quote_sheetname, absolute_coordinate
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import Workbook
from typing import Any, TypeVar, Dict, List
from openpyxl.styles import PatternFill, Font
from src.utils import get_github_token


DataFrame = TypeVar("DataFrame")
ExcelSheet = TypeVar("ExcelSheet")
ExcelWorkbook = TypeVar("ExcelWorkbook")


@dataclass
class ModelEndpoint:
    """Class for keeping track of model file endpoints"""

    model_file: str = (
        "https://raw.githubusercontent.com/CBIIT/ccdi-model/main/model-desc/ccdi-model.yml"
    )
    prop_file: str = (
        "https://raw.githubusercontent.com/CBIIT/ccdi-model/main/model-desc/ccdi-model-props.yml"
    )
    term_file: str = (
        "https://raw.githubusercontent.com/CBIIT/ccdi-model/main/model-desc/terms.yaml"
    )


@dataclass
class ManifestStyle:
    """Class for keeping track of style inventory"""

    # Metadata sheet style
    meta_linking_font: Any = Font(bold=True)
    meta_linking_pattern: Any = PatternFill(fill_type="solid", fgColor="DCD0FF")
    meta_index_font: Any = Font(bold=True)
    meta_index_pattern: Any = PatternFill(fill_type="solid", fgColor="DEFFF7")
    # Dictionary sheet style
    dict_header_pattern: Any = PatternFill(fill_type="solid", fgColor="000000")
    dict_header_font: Any = Font(bold=False, color="ffffff")
    # required prop style
    required_pattern: Any = PatternFill(fill_type="solid", fgColor="FFF2CC")
    required_font: Any = Font(bold=True)
    nonrequired_font: Any = Font(color="595959")
    # terms sheet pattern
    term_pattern_A: Any = PatternFill(fill_type="solid", fgColor="EEDDDC")
    term_pattern_B: Any = PatternFill(fill_type="solid", fgColor="DEE6F0")


class GetCCDIModel():
    node_preferred_order = [
        "study",
        "study_admin",
        "study_arm",
        "study_funding",
        "study_personnel",
        "publication",
        "participant",
        "family_relationship",
        "diagnosis",
        "therapeutic_procedure",
        "medical_history",
        "exposure",
        "radiology_file",
        "follow_up",
        "molecular_test",
        "sample",
        "cell_line",
        "pdx",
        "sequencing_file",
        "clinical_measure_file",
        "methylation_array_file",
        "cytogenomic_file",
        "pathology_file",
        "single_cell_sequencing_file",
        "synonym",
    ]

    def __init__(self, model_file: str, prop_file: str, term_file: str) -> None:
        self.model_file = model_file
        self.prop_file = prop_file
        self.term_file = term_file

    def _read_model(self) -> dict:
        model_dict = yaml.safe_load(Path(self.model_file).read_text())
        return model_dict

    def _read_prop(self) -> dict:
        prop_dict = yaml.safe_load(Path(self.prop_file).read_text())
        return prop_dict

    def _read_term(self) -> dict:
        term_dict = yaml.safe_load(Path(self.term_file).read_text())
        return term_dict

    def get_version(self) -> str:
        """Returns version value of data model"""
        model_dict = self._read_model()
        version = model_dict["Version"]
        return version

    def get_model_nodes(self) -> Dict:
        """Returns a dict of node and properties of the node"""
        model_dict = self._read_model()
        nodes_dict = model_dict["Nodes"]
        return nodes_dict

    def get_parent_node(self):
        """Gets parent nodes list of each node
        {
        "sample":["pdx","cell_line", "participant"],
        ...
        }
        """
        get_relations = self._read_model()["Relationships"]
        parent_node_dict = {}
        for i in get_relations.keys():
            child_node = i[3:]
            i_parent = []
            i_relation = get_relations[i]["Ends"]
            for h in i_relation:
                i_parent.append(h["Dst"])
            parent_node_dict[child_node] = i_parent
        return parent_node_dict

    def _read_each_prop(self, prop_dict: dict) -> tuple:
        """Extracts multiple prop information from a prop blob in a
        single prop dictionary

        Example prop_dict:
        {
            "Desc": "The text for reporting information about ethnicity based on the Office of Management and Budget (OMB) categories.",
            "Term": [
                {
                    "Origin": "caDSR",
                    "Code": "2192217",
                    "Value": "Ethnic Group Category Text",
                }
            ],
            "Type": {
                "value_type": "list",
                "Enum": [
                    "Hispanic or Latino",
                    "Not Allowed to Collect",
                    "Not Hispanic or Latino",
                    "Not Reported",
                    "Unknown",
                ],
            },
            "Req": True,
            "Strict": False,
            "Private": False,
        }
        """
        prop_description = prop_dict["Desc"]
        if "Term" in prop_dict.keys():
            term_list = prop_dict["Term"]
            cde_term = [i for i in term_list if i["Origin"] == "caDSR"]
            if len(cde_term) >= 1:
                prop_CDE = cde_term[0]["Code"]
            else:
                prop_CDE = np.nan
        else:
            prop_CDE = np.nan
        prop_required = prop_dict["Req"]
        if isinstance(prop_dict["Type"], str):
            # this covers string, integar, number
            prop_type = prop_dict["Type"]
        elif (
            prop_dict["Type"]["value_type"] == "string"
            and "Enum" in prop_dict["Type"].keys()
            and prop_dict["Strict"] == False
        ):
            # this covers string;enum
            prop_type = "string;enum"

        elif (
            prop_dict["Type"]["value_type"] == "string"
            and "Enum" in prop_dict["Type"].keys()
            and prop_dict["Strict"] == True
        ):
            # this covers enum
            prop_type = "enum"
        elif (
            prop_dict["Type"]["value_type"] == "list"
            and "Type" in prop_dict["Type"].keys()
        ):
            # this covers array[string]
            prop_type = "array[string]"
        elif (
            prop_dict["Type"]["value_type"] == "list"
            and "Enum" in prop_dict["Type"].keys()
            and prop_dict["Strict"] == True
        ):
            # this covers array[enum]
            prop_type = "array[enum]"
        elif (
            prop_dict["Type"]["value_type"] == "list"
            and "Enum" in prop_dict["Type"].keys()
            and prop_dict["Strict"] == False
        ):
            # this covers array[string;enum]
            prop_type = "array[string;enum]"
        else:
            print(json.dumps(prop_dict, indent=4))
            raise TypeError(
                "Can not categorize property type. Need to modify GetCCDIModel._read_each_prop method"
            )
        if isinstance(prop_dict["Type"], dict):
            if "Enum" in prop_dict["Type"].keys():
                prop_enum_list = prop_dict["Type"]["Enum"]
                if len(prop_enum_list) <= 4:
                    prop_example = ";".join(prop_enum_list)
                else:
                    prop_example = (
                        ";".join(prop_enum_list[0:4])
                        + ";etc (see Terms and Values Sets)"
                    )
            else:
                prop_example = np.nan
        else:
            prop_example = np.nan
        return prop_description, prop_type, prop_example, prop_required, prop_CDE

    def _get_prop_cde_version(self, prop_name: str, term_dict: dict) -> str:
        """Extracts CDE version of a property from a dict derived from terms.yaml

        Example blob of a prop term:
        {'Origin': 'caDSR',
        'Definition': 'The text for reporting information about ethnicity based on the Office of Management and Budget (OMB) categories.',
        'Code': 2192217,
        'Version': '2',
        'Value': 'Ethnic Group Category Text'}
        """
        if prop_name in term_dict.keys():
            # str() converts version into str type if number is found
            prop_cde_version = str(term_dict[prop_name]["Version"])
        else:
            prop_cde_version = np.nan
        return prop_cde_version

    def _get_sorted_node_list(self, node_list: list) -> list:
        """Sort the order of nodes and prioritize the
        preferred node order (self.node_preferred_order)
        """
        node_sort_list = self.node_preferred_order + [
            i for i in node_list if i not in self.node_preferred_order
        ]
        return node_sort_list

    def _if_enum_prop(self, prop_dict: dict) -> bool:
        if isinstance(prop_dict["Type"], dict):
            if "Enum" in prop_dict["Type"].keys():
                return True
            else:
                return False
        else:
            False

    def get_prop_dict_df(self):
        """Returns a dataframe that is ready to be loaded as "Dictionary" sheet"""
        # model_dict contains information of what properties in each node
        model_dict = self.get_model_nodes()
        # prop_dict contains property description, type, type, example value, and if the property is required
        prop_dict = self._read_prop()["PropDefinitions"]
        # term_dict contains CDE verson info
        term_dict = self._read_term()["Terms"]

        # create a dictionary to be convereted to df later
        prop_return_df = pd.DataFrame(
            columns=[
                "Property",
                "Description",
                "Node",
                "Type",
                "Example value",
                "Required",
                "Key",
                "CDE",
                "CDE version",
            ]
        )

        for node in model_dict.keys():
            node_property_list = model_dict[node]["Props"]
            for property in node_property_list:
                (
                    prop_description,
                    prop_type,
                    prop_example,
                    prop_required,
                    prop_cde,
                ) = self._read_each_prop(prop_dict=prop_dict[property])
                prop_cde_version = self._get_prop_cde_version(
                    prop_name=property, term_dict=term_dict
                )
                # change the prop required value to node name if True
                if prop_required == True:
                    prop_required = node
                else:
                    prop_required = np.nan
                # Mark Key property
                if node + "_id" == property:
                    prop_key = "TRUE"
                elif property == "id":
                    prop_key = "FALSE"
                else:
                    prop_key = np.nan
                prop_append_line = {
                    "Property": [property],
                    "Description": [prop_description],
                    "Node": [node],
                    "Type": [prop_type],
                    "Example value": [prop_example],
                    "Required": [prop_required],
                    "Key": [prop_key],
                    "CDE": [prop_cde],
                    "CDE version": [prop_cde_version],
                }
                prop_return_df = pd.concat(
                    [prop_return_df, pd.DataFrame(prop_append_line)], ignore_index=True
                )
        # sort the df based on node_preferred_order
        node_sort_list = self._get_sorted_node_list(node_list=model_dict.keys())
        prop_return_df.sort_values(
            by=["Node"],
            key=lambda column: column.map(lambda e: node_sort_list.index(e)),
            inplace=True,
        )
        return prop_return_df

    def get_terms_df(self):
        """Returns a dataframe that can be used for Terms and Value sets sheet"""
        # prop_dict contains property description, type, type, example value, and if the property is required
        prop_dict = self._read_prop()["PropDefinitions"]
        # term_dict contains CDE verson info
        term_dict = self._read_term()["Terms"]

        # Create an empty df
        terms_value_df = pd.DataFrame(
            columns=["Value Set Name", "(subset)", "Term", "Definition"]
        )

        # create a dictionary hosting df of each prop
        term_value_dict = {}

        # loop through every property in prop_dict
        for prop in prop_dict.keys():
            prop_context = prop_dict[prop]
            if self._if_enum_prop(prop_dict=prop_context):
                prop_enum_list = prop_context["Type"]["Enum"]
                term_definition_list = []
                for i in prop_enum_list:
                    if i in term_dict.keys():
                        term_definition_list.append(term_dict[i]["Definition"])
                    else:
                        term_definition_list.append(np.nan)
                prop_value_set_list = [prop] * len(prop_enum_list)
                subset_list = [np.nan] * len(prop_enum_list)
                prop_term_df = pd.DataFrame(
                    {
                        "Value Set Name": prop_value_set_list,
                        "(subset)": subset_list,
                        "Term": prop_enum_list,
                        "Definition": term_definition_list,
                    }
                )
                # add an empty row
                prop_term_df = pd.concat(
                    [
                        prop_term_df,
                        pd.DataFrame(
                            {
                                "Value Set Name": [np.nan],
                                "(subset)": [np.nan],
                                "Term": [np.nan],
                                "Definition": [np.nan],
                            }
                        ),
                    ],
                    ignore_index=True,
                )
                term_value_dict[prop] = prop_term_df
            else:
                pass
        # sort the key of dict term_Value_dict
        term_value_dict_keys = list(term_value_dict.keys())
        term_value_dict_keys.sort()
        term_value_dict = {i: term_value_dict[i] for i in term_value_dict_keys}
        # Concatenate prop_term_df to terms_value_df
        for prop in term_value_dict.keys():
            prop_term_df = term_value_dict[prop]
            terms_value_df = pd.concat(
                [terms_value_df, prop_term_df], ignore_index=True
            )
        return terms_value_df


class ManifestSheet():
    release_api = "https://api.github.com/repos/CBIIT/ccdi-model/releases"

    def __init__(self) -> None:
        self.workbook = Workbook()

    def _add_validation_to_sheet(
        self, ws: ExcelSheet, prop_name: str, excel_column_name: str
    ) -> None:
        """Adds data validation to the workbook sheet object"""
        validation_formula = "=" + prop_name
        dv = DataValidation(type="list", formula1=validation_formula, allow_blank=True)
        dv_range = excel_column_name + "2:" + excel_column_name + "1048576"
        ws.add_data_validation(dv)
        dv.add(dv_range)
        return None

    def _if_prop_required(
        self, prop_name: str, node_name: str, prop_dict_df: DataFrame, logger
    ) -> bool:
        prop_required_value = prop_dict_df.loc[
            (prop_dict_df["Property"] == prop_name)
            & (prop_dict_df["Node"] == node_name),
            "Required",
        ].values
        if len(prop_required_value) > 1:
            logger.error(
                f"Property {prop_name} appears more than once in Dictionary sheet"
            )
            if prop_required_value[0] == node_name:
                return True
            else:
                return False
        else:
            if prop_required_value[0] == node_name:
                return True
            else:
                return False

    def _get_sheets_order(self, expected_name_order: List) -> None:
        """Returns a list of index which helps to reoder sheets in
        workbook

        Example: reorder sheets ["sheetA", "sheetB", "sheetC"] into ["sheetC","sheetA", "sheetB"]
        return would be [2, 0, 1]
        """
        sheetnames = self.workbook.sheetnames
        reorder_index = []
        for i in expected_name_order:
            i_expected_index = sheetnames.index(i)
            reorder_index.append(i_expected_index)
        return reorder_index

    def release_api_return(self) -> List[Dict]:
        """Get a return from github repo ccdi-model release api"""
        github_token =  get_github_token()
        headers = {"Authorization": "token " + github_token}
        release_endpoint_response = requests.get(self.release_api, headers=headers)
        response_list = release_endpoint_response.json()
        version_list = []
        title_list = []
        tag_url_list = []
        for i in response_list:
            i_version, i_title = i["name"].split(":")
            i_title = i_title.strip()
            i_tag_url = i["html_url"]
            version_list.append(i_version)
            title_list.append(i_title)
            tag_url_list.append(i_tag_url)
        version_list.reverse()
        title_list.reverse()
        tag_url_list.reverse()
        return version_list, title_list, tag_url_list

    def dictionary_sheet(self, dict_df: DataFrame) -> None:
        """Creates Dictionary sheet"""
        sheet_dictionary = self.workbook.create_sheet(title="Dictionary")
        # write each row to sheet "Dictionary"
        for r in dataframe_to_rows(dict_df, index=False, header=True):
            sheet_dictionary.append(r)
        # format header pattern and font in sheet "Dictionary"
        for h in [get_column_letter(i + 1) for i in range(dict_df.shape[1])]:
            cell_pos = h + "1"
            sheet_dictionary[cell_pos].fill = ManifestStyle.dict_header_pattern
            sheet_dictionary[cell_pos].font = ManifestStyle.dict_header_font
            sheet_dictionary.column_dimensions[h].width = 25
        # format required and nonrequired property cell
        row_number = 1
        for row in dataframe_to_rows(dict_df, index=False, header=True):
            # if prop is id, file_url_in_cds or dcf_indexd_guid
            if row[0] in ["id", "file_url_in_cds", "dcf_indexd_guid"]:
                sheet_dictionary["A" + str(row_number)].font = (
                    ManifestStyle.meta_index_font
                )
                sheet_dictionary["A" + str(row_number)].fill = (
                    ManifestStyle.meta_index_pattern
                )
            else:
                if pd.isna(row[5]):
                    sheet_dictionary["A" + str(row_number)].font = (
                        ManifestStyle.nonrequired_font
                    )
                else:
                    sheet_dictionary["A" + str(row_number)].font = (
                        ManifestStyle.required_font
                    )
                    sheet_dictionary["A" + str(row_number)].fill = (
                        ManifestStyle.required_pattern
                    )
            row_number += 1
        # freeze first row of dictionary sheet
        sheet_dictionary.freeze_panes = "A2"
        return None

    def readme_sheet(self, model_version: str, release_title: str) -> None:
        """Creates README and INSTRUCTIONS sheet
        """
        top_df_dict = {
            "col1": [
                "CCDI SUBMISSION METADATA TEMPLATE",
                np.nan,
                "OVERVIEW",
                np.nan,
                np.nan,
                "INSTRUCTIONS FOR SUBMISSION METADATA TEMPLATE",
                np.nan,
                np.nan,
                "STRUCTURED COLUMNS",
                np.nan,
                np.nan,
                np.nan,
                np.nan,
                "REQUIRED DATA",
                np.nan,
                np.nan,
                np.nan,
                np.nan,
                "OTHER REQUIRED DATA",
                np.nan,
                "REQUIRED DATA FOR DATA FILES",
                np.nan,
                "DICTIONARY, TERMS AND VALUE SETS",
                "STRUCTURED COLUMNS ",
                np.nan,
                np.nan,
                np.nan,
                np.nan,
                "NEED HELP? HAVE A QUESTION? HAVE FEEDBACK?",
                np.nan,
                np.nan,
                "VERSION",
            ],
            "col2": [
                np.nan,
                np.nan,
                np.nan,
                "This is the metadata template for submitting data to the Childhood Caner Data Initiative (CCDI)",
                "The submission metadata is useful to help make data FAIR for the potential users.",
                np.nan,
                "Please do not delete columns, please keep the columns in the same order.",
                "Please fill out the required fields (see below).",
                np.nan,
                "The structured fields are denoted by BOLD black font with a purple background.",
                "The first column, type, is a value that notes the tab you are in. Please do not replace with other text and make sure each row has the value for that node.",
                "When linking data from one tab to another, if there are multiple linking columns [node.node_id], please try to only link to one node.",
                "While linking can happen at multiple levels, it is best practice to link it to the lowest level as the parent node should eventually create that same linkage.",
                np.nan,
                "The required fields are denoted by BOLD black font with a yellow background.",
                'These are also marked as being required on the dictionary tab with labels based on the section of data, like "study", "sample" or "participant".',
                "Optional fields are marked with grey font",
                "Each of the fields on are described in Dictionary, along with terms and value sets (enumerated values).",
                np.nan,
                "Properties that are highlighted in blue are required in the model but are not required from the user at the time of submission.",
                np.nan,
                "For sequencing files, please try to provide all metadata, if applicable, for the following properties: avg_read_length, number_of_reads, number_of_bp, coverage",
                np.nan,
                'The "Dictionary" defines the top-level parts (or categories or tabs), their fields and allowed values.',
                'Note that "Value Sets" column describes the name of the value set of enumerated values. The value sets and comprised terms are found on the tab "Terms and Value Sets". ',
                "Note that the column CDE provides the code used for the value set for reference purposes.",
                "The required fields are denoted with yellow background with black bold font. While not all fields are required, it is best practice to supply as much data as possible.",
                'To see the expanded set of values in the Terms and Value Sets tab, click the "plus" button on the left side of the table.',
                np.nan,
                "Feel free to reach out if you have questions or need help filling out this metadata manifest for your submission!",
                "Contact us at CCDIHelpDesk@mail.nih.gov",
                np.nan,
            ],
            "col3": [model_version] + [np.nan] * 31,
            "col4": [np.nan] * 32,
        }

        # the bottom df for the release history info
        verion_list, title_list, tag_url_list = self.release_api_return()
        verion_list.append(model_version)
        title_list.append(release_title)
        tag_url_list.append(
            "https://github.com/CBIIT/ccdi-model/releases/tag/" + model_version[1:]
        )
        bottom_df_dict = {
            "col1": [np.nan] * len(verion_list),
            "col2": verion_list,
            "col3": title_list,
            "col4": tag_url_list,
        }

        # Concatenate top and bottom dfs
        readme_df = pd.concat(
            [pd.DataFrame(top_df_dict), pd.DataFrame(bottom_df_dict)], ignore_index=True
        )

        # write the combined df to a workbook sheet
        sheet_readme = self.workbook.create_sheet(title="README and INSTRUCTIONS")
        for r in dataframe_to_rows(readme_df, index=False, header=False):
            sheet_readme.append(r)
        # freeze first row of dictionary sheet
        sheet_readme.freeze_panes = "A2"

        # adjust column width
        for h in [get_column_letter(i + 1) for i in range(4)]:
            sheet_readme.column_dimensions[h].width = 25
        return None

    def terms_value_sets_sheet(self, terms_df: DataFrame) -> None:
        """Create Terms and Value Sets sheet"""
        # write the combined df to a workbook sheet
        sheet_terms = self.workbook.create_sheet(title="Terms and Value Sets")
        for r in dataframe_to_rows(terms_df, index=False, header=True):
            sheet_terms.append(r)

        # adjust column width and header font, patternfill
        for h in [get_column_letter(i + 1) for i in range(4)]:
            sheet_terms.column_dimensions[h].width = 25
            sheet_terms[h + "1"].fill = ManifestStyle.dict_header_pattern
            sheet_terms[h + "1"].font = ManifestStyle.dict_header_font
        sheet_terms.column_dimensions["D"].width = 200

        # collapse rows for each value set name
        for i in terms_df["Value Set Name"].unique().tolist():
            if not pd.isna(i):
                i_index = terms_df[terms_df["Value Set Name"] == i].index.tolist()
                i_start_row = i_index[0] + 3
                i_end_row = i_index[-1] + 2
                sheet_terms.row_dimensions.group(i_start_row, i_end_row, hidden=True)
            else:
                pass

        # adjust background style
        bgcolor_rotate = 1
        for i in terms_df["Value Set Name"].unique().tolist():
            if not pd.isna(i):
                i_index = terms_df[terms_df["Value Set Name"] == i].index.tolist()
                i_start_row = i_index[0] + 2
                i_end_row = i_index[-1] + 4
                if (bgcolor_rotate % 2) == 1:
                    fill_style = ManifestStyle.term_pattern_A
                else:
                    fill_style = ManifestStyle.term_pattern_B
                for h in range(i_start_row, i_end_row):
                    for k in ["A", "B", "C", "D"]:
                        ind_cell = k + str(h)
                        sheet_terms[ind_cell].fill = fill_style
                    # h_row =  sheet_terms.row_dimensions[h]
                    # h_row.fill = fill_style
                bgcolor_rotate += 1
            else:
                pass

        return None

    def get_define_names(self, term_df: DataFrame) -> None:
        """Create defined names for the workbook
        This method can only be done after terms and value sets sheet
        has been created and before the creation of node metadata sheet
        """
        # get worksheet
        ws = self.workbook["Terms and Value Sets"]
        # get unique value set names
        value_set_names = term_df["Value Set Name"].dropna().unique().tolist()
        # loop through item in value_set_names and add defined names to workbook
        for i in value_set_names:
            i_index_list = term_df[term_df["Value Set Name"] == i].index.tolist()
            i_begin_row = i_index_list[0]
            i_end_row = i_index_list[-1]
            i_range = f"C{i_begin_row+2}:C{i_end_row+2}"
            i_defined_name_formula = (
                f"{quote_sheetname(ws.title)}!{absolute_coordinate(i_range)}"
            )
            i_defined_name_object = DefinedName(i, attr_text=i_defined_name_formula)
            self.workbook.defined_names[i] = i_defined_name_object
        return None

    def node_metadata_sheet(
        self,
        node: str,
        model_node: Dict,
        parent_node_dict: Dict,
        prop_dict_df: DataFrame,
        logger,
    ) -> None:
        """Create a single metadata sheet for a node"""
        # create a metadata sheet for node
        self.workbook.create_sheet(node)
        logger.info(f"Creating metadata sheet for node {node}")
        ws_node = self.workbook[node]
        # create metadata sheet header
        if node != "study":
            parent_nodes = parent_node_dict[node]
            parent_nodes_extended = [i + "." + i + "_id" for i in parent_nodes]
            parent_nodes_index = [i + ".id" for i in parent_nodes]
        else:
            parent_nodes = []
            parent_nodes_extended = []
            parent_nodes_index = []

        node_props = model_node[node]["Props"][:-1]
        node_sheet_header = (
            ["type"] + parent_nodes_extended + node_props + ["id"] + parent_nodes_index
        )
        # add node_sheet_header as metadata sheet header
        ws_node.append(node_sheet_header)
        ws_node.append([node])

        # add styles to the header
        # type and linking properties
        cols_linking = [
            get_column_letter(i + 1)
            for i in range(len(["type"] + parent_nodes_extended))
        ]
        for col in cols_linking:
            ws_node.column_dimensions[col].width = 25
            cell_col = col + "1"
            ws_node[cell_col].fill = ManifestStyle.meta_linking_pattern
            ws_node[cell_col].font = ManifestStyle.meta_linking_font
        # node props
        for h in range(len(node_props)):
            h_col_name = get_column_letter(
                h + len(["type"] + parent_nodes_extended) + 1
            )
            ws_node.column_dimensions[h_col_name].width = 25
            cell_h_col = h_col_name + "1"
            h_prop = node_props[h]
            # add data validation for column if h_prop has been defined
            if h_prop in self.workbook.defined_names.keys():
                self._add_validation_to_sheet(
                    ws=ws_node, prop_name=h_prop, excel_column_name=h_col_name
                )
            else:
                pass
            # Add style to required or non-required property column
            if_h_prop_req = self._if_prop_required(
                prop_name=h_prop,
                node_name=node,
                prop_dict_df=prop_dict_df,
                logger=logger,
            )
            if if_h_prop_req:
                if h_prop in ["file_url_in_cds", "dcf_indexd_guid"]:
                    ws_node[cell_h_col].fill = ManifestStyle.meta_index_pattern
                    ws_node[cell_h_col].font = ManifestStyle.meta_index_font
                else:
                    ws_node[cell_h_col].fill = ManifestStyle.required_pattern
                    ws_node[cell_h_col].font = ManifestStyle.required_font
            else:
                ws_node[cell_h_col].font = ManifestStyle.nonrequired_font
        # index props
        for k in range(len(["id"] + parent_nodes_index)):
            k_col_name = get_column_letter(
                k + 1 + len(["type"] + parent_nodes_extended + node_props)
            )
            ws_node.column_dimensions[k_col_name].width = 25
            cell_k_col = k_col_name + "1"
            ws_node[cell_k_col].fill = ManifestStyle.meta_index_pattern
            ws_node[cell_k_col].font = ManifestStyle.meta_index_font
        # freeze first row
        ws_node.freeze_panes = "A2"
        return None

    def metadata_sheets(
        self, model_node: Dict, parent_node_dict: Dict, prop_dict_df, logger
    ):
        """Create metadata sheet for every node in the workbook"""
        for node in model_node.keys():
            self.node_metadata_sheet(
                node=node,
                model_node=model_node,
                parent_node_dict=parent_node_dict,
                prop_dict_df=prop_dict_df,
                logger=logger,
            )
        return None

    def sort_sheets(self, sorted_node_list: List) -> None:
        """Sorts sheets order in the workbook
        This should be fone after all sheets have been created
        """
        del self.workbook["Sheet"]
        sheets_order = (
            ["README and INSTRUCTIONS"]
            + sorted_node_list
            + ["Dictionary", "Terms and Value Sets"]
        )
        sort_sheet_index = self._get_sheets_order(expected_name_order=sheets_order)
        self.workbook._sheets = [self.workbook._sheets[i] for i in sort_sheet_index]
