import mock
import sys
import os
import pytest
import pandas as pd
import random
from openpyxl import Workbook
from pathlib import Path

parent_dir = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
sys.path.append(parent_dir)
from src.create_submission import GetCCDIModel, ManifestSheet
import logging
import numpy as np
import bento_meta


@pytest.fixture
def getccdimodel_object():
    """GetCCDIModel object for testing purpose"""
    model_yml = "./tests/test_files/test-ccdi-model.yml"
    props_yml = "./tests/test_files/test-ccdi-model-props.yml"
    terms_yml = "./tests/test_files/test-terms.yml"
    obj = GetCCDIModel(
        model_file=model_yml, prop_file=props_yml, term_file=terms_yml
    )
    return obj


@pytest.fixture
def getmanifestsheet_object():
    """Get ManifestSheet object for testing purpose"""
    obj = ManifestSheet()
    return obj


def test_getccdimodel_init(getccdimodel_object):
    """test getccdimodel init"""
    assert os.path.basename(getccdimodel_object.model_file) == "test-ccdi-model.yml"
    assert os.path.basename(getccdimodel_object.prop_file) == "test-ccdi-model-props.yml"
    assert (
        os.path.basename(getccdimodel_object.term_file) == "test-terms.yml"
    )
    assert isinstance(getccdimodel_object.ccdi_model, bento_meta.model.Model)

def test_getccdimodel_version(getccdimodel_object):
    """test get version of model"""
    assert getccdimodel_object.get_version() == "v1.9.1"

def test_getccdimodel_list_node_props(getccdimodel_object):
    """test _list_node_props"""
    study_node_props =  getccdimodel_object._list_node_props(node_name = "study")
    assert len(study_node_props) == 13
    assert "study_acronym" in study_node_props

def test_getccdimodel_parent_nodes(getccdimodel_object):
    parent_nodes_dict = getccdimodel_object.get_parent_nodes()
    assert parent_nodes_dict["sample"] == ["participant", "pdx", "cell_line"]
    assert parent_nodes_dict["study_arm"] == ["study"]

def test_getccdimodel_list_nodes(getccdimodel_object):
    node_list = getccdimodel_object._list_nodes()
    assert len(node_list) == 25
    assert "sequencing_file" in node_list
    assert "single_cell_seqeuncing_file" not in node_list

def test_getccdimodel_get_prop_cde_code(getccdimodel_object):
    prop_obj = getccdimodel_object.ccdi_model.nodes["radiology_file"].props[
        "scanner_manufacturer"
    ]
    assert getccdimodel_object._get_prop_cde_code(prop_obj=prop_obj) == "2866141"
    assert isinstance(prop_obj, bento_meta.objects.Property)

def test_getccdimodel_read_each_prop(getccdimodel_object):
    enum_strict_prop = getccdimodel_object._read_each_prop(
        node_name="survival", prop_name="last_known_survival_status"
    )
    enum_nonstrict_prop = getccdimodel_object._read_each_prop(
        node_name="survival", prop_name="follow_up_category"
    )
    list_enum_prop = getccdimodel_object._read_each_prop(
        node_name="study", prop_name="study_data_types"
    )
    list_enum_prop_nonstrict = getccdimodel_object._read_each_prop(
        node_name="diagnosis", prop_name="anatomic_site"
    )
    integer_prop = getccdimodel_object._read_each_prop(
        node_name="survival", prop_name="age_at_last_known_survival_status"
    )
    number_prop = getccdimodel_object._read_each_prop(
        node_name="exposure", prop_name="alcohol_drinks_per_day"
    )
    string_prop = getccdimodel_object._read_each_prop(
        node_name = "sequencing_file", prop_name="md5sum"
    )
    key_prop = getccdimodel_object._read_each_prop(
        node_name="study", prop_name="study_id"
    )
    assert enum_strict_prop[1] == "enum"
    assert enum_nonstrict_prop[1] == "string;enum"
    assert list_enum_prop[1] == "array[enum]"
    assert list_enum_prop_nonstrict[1] == "array[string;enum]"
    assert integer_prop[1] == "integer"
    assert number_prop[1] == "number"
    assert string_prop[1] == "string"
    assert string_prop[3] # test if required
    assert key_prop[4] # test if key
    assert not number_prop[3] # test not required
    assert string_prop[4] is np.nan # test if key, expect nan

def test_get_sorted_node_list(getccdimodel_object):
    preferred_order = getccdimodel_object.node_preferred_order
    test_node_list = random.sample(preferred_order, len(preferred_order))
    test_node_list_addition = random.sample(preferred_order, len(preferred_order)) + [
        "nodeA",
        "nodeB",
    ]
    assert (
        getccdimodel_object._get_sorted_node_list(test_node_list)
        == getccdimodel_object.node_preferred_order
    )
    assert (
        getccdimodel_object._get_sorted_node_list(test_node_list_addition)[:-2]
        == getccdimodel_object.node_preferred_order
    )


def test_manifestsheet_init(getmanifestsheet_object):
    workbook = getmanifestsheet_object.workbook
    assert isinstance(workbook, Workbook)
    assert (
        getmanifestsheet_object.release_api
        == "https://api.github.com/repos/CBIIT/ccdi-model/releases"
    )


def test_manifestsheet_get_sheets_order(getmanifestsheet_object):
    getmanifestsheet_object.workbook.create_sheet("sheetA")
    getmanifestsheet_object.workbook.create_sheet("sheetB")
    getmanifestsheet_object.workbook.create_sheet("sheetC")
    del getmanifestsheet_object.workbook["Sheet"]
    print(getmanifestsheet_object.workbook.sheetnames)
    order_index = getmanifestsheet_object._get_sheets_order(
        expected_name_order=["sheetC", "sheetB", "sheetA"]
    )
    assert order_index == [2, 1, 0]


def test_manifestsheet_if_prop_required(getmanifestsheet_object):
    prop_dict_test = pd.DataFrame(
        {
            "Property": ["propertyA", "propertyB", "propertyC"],
            "Node": ["Node1", "Node2", "Node3"],
            "Required": [np.nan, "Node2", "Node3"],
        }
    )
    if_propA_required = getmanifestsheet_object._if_prop_required(
        prop_name="propertyA", node_name="Node1", prop_dict_df=prop_dict_test, logger=logging
    )
    if_propB_required = getmanifestsheet_object._if_prop_required(
        prop_name="propertyB", node_name="Node2", prop_dict_df=prop_dict_test, logger=logging
    )
    assert not if_propA_required
    assert if_propB_required


@mock.patch("src.create_submission.requests", autospec=True)
def test_manifestsheet_release_api_return(mock_requests, getmanifestsheet_object):
    release_data_test = [
        {
            "name": "v0.1.1: test title 2",
            "html_url": "https://github.com/test/test_repo/releases/tag/0.1.1",
        },
        {
            "name": "v0.1.0: test title 1",
            "html_url": "https://github.com/test/test_repo/releases/tag/0.1.0",
        },
        
    ]
    mock_requests_get =  mock_requests.get.return_value
    mock_requests_get.json.return_value = release_data_test
    version_list, title_list, tag_url_list = getmanifestsheet_object.release_api_return()
    assert version_list == ["v0.1.0","v0.1.1"]
    assert title_list == ["test title 1", "test title 2"]
    assert tag_url_list == [
        "https://github.com/test/test_repo/releases/tag/0.1.0",
        "https://github.com/test/test_repo/releases/tag/0.1.1",
    ]

