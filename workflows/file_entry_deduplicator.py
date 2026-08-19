import pandas as pd
import re
import os
import shutil
import sys
import copy
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from prefect import task, flow, get_run_logger
from src import get_time, file_dl, folder_ul

timestamp = get_time()

def find_file_tabs(excel_path: str) -> list[str]:
    """Find all tabs that match the pattern [node]_file."""
    xl = pd.ExcelFile(excel_path)
    return [sheet for sheet in xl.sheet_names if sheet.endswith("_file")]


def get_node_id_col(tab_name: str, df: pd.DataFrame) -> str | None:
    """Find the [node]_file_id column for a given tab."""
    expected = f"{tab_name}_id"
    matches = [c for c in df.columns if c == expected]
    return matches[0] if matches else None


def get_parent_cols(df: pd.DataFrame) -> list[str]:
    """Find all parent linking columns in the format [node].[node]_id."""
    return [c for c in df.columns if re.match(r"^\w+\.\w+_id$", c)]


def validate_non_parent_cols(
    group: pd.DataFrame,
    dedup_keys: list[str],
    node_id_col: str,
    parent_cols: list[str],
) -> dict:
    """
    For a group of duplicate rows, validate that all non-key, non-parent,
    non-id columns have the same value across all rows.

    Returns a dict of {col: list_of_conflicting_values} for any conflicts.
    """
    exclude = set(dedup_keys + parent_cols + [node_id_col])
    conflicts = {}
    for col in group.columns:
        if col in exclude:
            continue
        unique_vals = group[col].dropna().unique()
        if len(unique_vals) > 1:
            conflicts[col] = unique_vals.tolist()
    return conflicts


def resolve_node_id(
    group: pd.DataFrame,
    node_id_col: str,
    file_name_col: str = "file_name",
) -> tuple[str, list]:
    """
    Determine the merged [node]_file_id value.
    - If all node_id values are just variations of file_name, use file_name.
    - Otherwise return a list of all unique ids and flag for review.

    Returns (resolved_id, warnings)
    """
    warnings = []
    id_values = group[node_id_col].dropna().unique().tolist()

    if file_name_col in group.columns:
        file_name_vals = group[file_name_col].dropna().unique()
        file_name = file_name_vals[0] if len(file_name_vals) == 1 else None

        if file_name:
            # check if all ids are just variations of the file_name
            all_variations = all(
                file_name in str(id_val) or str(id_val) in file_name
                for id_val in id_values
            )
            if all_variations:
                return file_name, warnings

    # ids are not simple file_name variations — return list and warn
    warnings.append(
        f"Could not resolve {node_id_col} to file_name. "
        f"Unique IDs found: {id_values}. Manual review required."
    )
    return ";".join(str(i) for i in id_values), warnings

@task
def merge_duplicate_file_rows(
    df: pd.DataFrame,
    tab_name: str,
    dedup_keys: list[str] = None,
) -> tuple[pd.DataFrame, list[str]]:
    """
    Merge duplicate file rows based on dedup_keys (default: md5sum, file_url, dcf_indexd_guid).
    Parent columns are concatenated with ';'.
    Non-parent, non-key columns must be identical across duplicates or a conflict is flagged.

    Returns (merged_df, list_of_warnings)
    """
    logger = get_run_logger()

    if dedup_keys is None:
        dedup_keys = ["md5sum", "file_url", "dcf_indexd_guid"]

    # only use dedup_keys that exist in the df
    active_dedup_keys = [k for k in dedup_keys if k in df.columns]
    if not active_dedup_keys:
        return df, [f"[{tab_name}] None of the dedup keys {dedup_keys} found in columns, skipping."]

    node_id_col = get_node_id_col(tab_name, df)
    if not node_id_col:
        return df, [f"[{tab_name}] Could not find node id column, skipping."]

    parent_cols = get_parent_cols(df)
    all_warnings = []
    merged_rows = []

    # separate duplicate groups from unique rows
    duplicated_mask = df.duplicated(subset=active_dedup_keys, keep=False)
    df_unique = df[~duplicated_mask].copy()
    df_dupes = df[duplicated_mask].copy()

    if df_dupes.empty:
        logger.info(f"[{tab_name}] No duplicate rows found.")
        return df, []

    logger.info(f"[{tab_name}] Found {len(df_dupes)} duplicate rows across "
        f"{df_dupes.groupby(active_dedup_keys).ngroups} groups.")

    for key_vals, group in df_dupes.groupby(active_dedup_keys, dropna=False):
        # ── validation: check non-key, non-parent, non-id cols are identical ──
        conflicts = validate_non_parent_cols(
            group=group,
            dedup_keys=active_dedup_keys,
            node_id_col=node_id_col,
            parent_cols=parent_cols,
        )
        if conflicts:
            conflict_summary = "; ".join(
                f"{col}={vals}" for col, vals in conflicts.items()
            )
            all_warnings.append(
                f"[{tab_name}] Conflict in duplicate group "
                f"{dict(zip(active_dedup_keys, key_vals if isinstance(key_vals, tuple) else [key_vals]))}: "
                f"{conflict_summary}. Rows kept separate."
            )
            # keep rows as-is if there are conflicts
            for _, row in group.iterrows():
                merged_rows.append(row.to_dict())
            continue

        # ── merge parent columns with ';' ─────────────────────────────────────
        merged_row = group.iloc[0].to_dict()
        for parent_col in parent_cols:
            unique_parents = group[parent_col].dropna().unique().tolist()
            # flatten any already-semicolon-joined values
            flat_parents = []
            for p in unique_parents:
                for part in str(p).split(";"):
                    part = part.strip()
                    if part and part not in flat_parents:
                        flat_parents.append(part)
            merged_row[parent_col] = ";".join(flat_parents) if flat_parents else None

        # ── resolve node id ───────────────────────────────────────────────────
        resolved_id, id_warnings = resolve_node_id(
            group=group,
            node_id_col=node_id_col,
        )
        if id_warnings:
            all_warnings.extend([f"[{tab_name}] {w}" for w in id_warnings])
        merged_row[node_id_col] = resolved_id

        merged_rows.append(merged_row)

    # combine merged dupes with unique rows and restore original column order
    df_merged_dupes = pd.DataFrame(merged_rows, columns=df.columns)
    df_final = pd.concat([df_unique, df_merged_dupes], ignore_index=True)
    df_final = df_final[df.columns]  # restore original column order

    logger.info(f"[{tab_name}] Reduced {len(df_dupes)} duplicate rows to "
        f"{len(df_merged_dupes)} merged rows.")
    return df_final, all_warnings


def write_sheets_preserve_formatting(
    all_sheets: dict, input_file: str, output_file: str
) -> None:
    """
    Write updated dataframes back to the workbook starting at row 2,
    leaving the formatted header row 1 completely untouched.
    """
    logger = get_run_logger()

    wb = load_workbook(input_file)

    for sheet_name, df in all_sheets.items():
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            continue

        ws = wb[sheet_name]

        # clear existing data rows only (row 2 onwards)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        # write new data starting at row 2, skipping the header
        for row_idx, row_data in enumerate(
            dataframe_to_rows(df, index=False, header=False), start=2
        ):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx).value = value

    wb.save(output_file)
    logger.info(f"Saved with formatting preserved: {output_file}")


@flow(name="Manifest Deduplicator",
        log_prints=True,
        description="Deduplicate *_file tabs in a manifest Excel file.",
        flow_run_name="{runner}_" + f"{timestamp}",
        )
def process_manifest(
    bucket: str, 
    runner: str,
    file_path: str, 
    ) -> None:
    """A process to deduplicate *_file tabs in a manifest Excel file by loading a manifest, find all _file tabs, merge duplicate rows, and save the result with all formatting preserved.

    Args:
        bucket (str): The bucket where the manifest is stored.
        runner (str): The runner executing the flow.
        file_path (str): Path to the input manifest Excel file.
    """
    logger = get_run_logger()
    logger.info(f"Starting manifest deduplication for file: {file_path}")

    logger.info(f"Downloading file from bucket={bucket}, path={file_path}")
    file_dl(bucket=bucket, file_path=file_path)

    file_name = Path(file_path).name
    output_file = f"{file_name}_deduped_{timestamp}.xlsx"
    output_file_log = f"{file_name}_deduped_{timestamp}.log"

    file_tabs = find_file_tabs(file_name)
    if not file_tabs:
        logger.info("No *_file tabs found in the manifest.")
        return

    logger.info(f"Found file tabs: {file_tabs}")

    all_warnings = []

    logger.info(f"Reading Excel file: {file_name}")
    # read all sheets up front
    all_sheets = pd.read_excel(file_name, sheet_name=None, dtype=str)

    for tab in file_tabs:
        logger.info(f"\nProcessing tab: {tab}")
        df = all_sheets[tab]

        df_merged, warnings = merge_duplicate_file_rows(df=df, tab_name=tab)
        all_sheets[tab] = df_merged
        all_warnings.extend(warnings)

    logger.info(f"\nDeduplication complete. Writing output to: {output_file}")
    # write all sheets back with formatting preserved
    write_sheets_preserve_formatting(
        all_sheets=all_sheets,
        input_file=file_name,
        output_file=output_file,
    )

    logger.info(f"Writing warnings log to: {output_file_log}")
    with open(output_file_log, "w") as log_file:
        if all_warnings:
            print(f"\n{'='*60}", file=log_file)
            print("WARNINGS — Manual review required:", file=log_file)
            print("="*60, file=log_file)
            for w in all_warnings:
                print(f"  ⚠  {w}", file=log_file)
        else:
            print("\nNo warnings — all duplicates merged cleanly.", file=log_file)

    logger.info(f"Deduplication process completed. Output files: {output_file}, {output_file_log}")
    logger.info(f"Moving output files to bucket={bucket}, destination={runner}/{timestamp}")
    # Create an output directory for the deduplicated files and then move the files there
    output_directory = os.path.join(f"deduplicated_{timestamp}")
    os.makedirs(output_directory, exist_ok=True)
    shutil.move(file_name, output_directory)
    shutil.move(output_file, output_directory)
    shutil.move(output_file_log, output_directory)

    folder_ul(local_folder=output_directory, bucket=bucket, destination=runner, sub_folder="")

    logger.info(f"Deduplication complete. Deduplicated files and logs moved to bucket={bucket}, destination={runner}/{output_directory}")
