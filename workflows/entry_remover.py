#!/usr/bin/env python3
"""
entry_remover.py

Remove specified entries (and any linked “child” entries) from a CCDI metadata manifest.
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from prefect import get_run_logger, task, flow
from yaml import warnings
from src.utils import file_dl, get_time, file_ul, folder_ul, folder_dl


# @task(name="Drop empty rows/columns")
# def drop_empty(df, axis):
#     """
#     Remove rows (axis=0) or columns (axis=1) that are entirely NA/blank.
#     """
#     return df.dropna(how="all", axis=axis)


@flow(name="Remover")
def main(file, directory, entry):

    logger = get_run_logger()
    logger.info("Starting entry removal process")

    entry_path = entry

    if file:
        logger.info(f"Using manifest file: {file}")
        # Absolute paths & output names
        manifest_path = file

        base = os.path.splitext(os.path.basename(manifest_path))[0]
        today = get_time()
        out_xlsx = f"{base}_EntRemove_{today}.xlsx"
        out_delete_xlsx = f"{base}_EntRemove_{today}_deleted.xlsx"
        log_txt = f"{base}_EntRemove_{today}_log.txt"
        out_folder = None  # not used when input is a single file

        ##############
        #
        # Pull Dictionary Page to create node pulls
        #
        ##############

        logger.info("Reading CCDI Manifest file")

        # create workbook
        xlsx_data = pd.ExcelFile(manifest_path)

        # create dictionary for dfs
        meta_dfs = {}

        # 1) Read in the manifest and create dfs for each tab in the manifest
        for sheet in xlsx_data.sheet_names:
            meta_dfs[sheet] = pd.read_excel(
                manifest_path,
                sheet_name=sheet,
                dtype=str,
                na_values=["NA", "na", "N/A", "n/a"],
                keep_default_na=False,
                engine="openpyxl",
            )

        # remove model tabs from the meta_dfs
        del meta_dfs["README and INSTRUCTIONS"]
        del meta_dfs["Dictionary"]
        del meta_dfs["Terms and Value Sets"]

        # create a list of present tabs
        dict_nodes = set(list(meta_dfs.keys()))

        # Go through each tab and remove completely empty tabs
        logger.info("Removing empty tabs from the manifest file")

        for node in dict_nodes:
            # see if the tab contain any data
            test_df = meta_dfs[node]
            test_df = test_df.drop("type", axis=1)
            test_df = test_df.dropna(how="all").dropna(how="all", axis=1)
            # if there is no data, drop the node/tab
            if test_df.empty:
                del meta_dfs[node]

        # # determine nodes again
        # node_list = set(list(meta_dfs.keys()))

        # create a blank copy of the meta_dfs to save deleted entries to, and to modify for output
        meta_dfs_delete = {
            node: pd.DataFrame(columns=df.columns) for node, df in meta_dfs.items()
        }

    if directory:
        logger.info(f"Using manifest directory: {directory}")
        # Absolute paths & output names
        manifest_path = directory
        
        base = os.path.basename(manifest_path)
        today = get_time()
        out_folder = f"{base}_EntRemove_{today}"
        base_out_tsv = f"_EntRemove_{today}.tsv"
        out_delete_xlsx = f"{base}_EntRemove_{today}_deleted.xlsx"
        log_txt = f"{base}_EntRemove_{today}_log.txt"
        out_xlsx = None  # not used when input is a directory of TSVs

        # read in all tsvs in the directory and create dfs for each
        meta_dfs = {}
        for filename in os.listdir(manifest_path):
            if filename.endswith(".tsv"):
                # the node name will be determined by reading the file and using the first row of the "type" column, which should be the node name. This is more robust than using the filename, which may not always be consistent.
                temp_df = pd.read_csv(
                    os.path.join(manifest_path, filename),
                    sep="\t",
                    dtype=str,
                    na_values=["NA", "na", "N/A", "n/a"],
                    keep_default_na=False,
                )
                node_name = temp_df.loc[0, "type"]
                meta_dfs[node_name] = pd.read_csv(
                    os.path.join(manifest_path, filename),
                    sep="\t",
                    dtype=str,
                    na_values=["NA", "na", "N/A", "n/a"],
                    keep_default_na=False,
                )

    # 3) Read entries to remove
    logger.info(f"Reading entries to remove from {entry_path}")
    entries_df = pd.read_csv(entry_path, sep="\t", header=None, names=["X1"], dtype=str)
    pending = entries_df["X1"].dropna().tolist()

    deleted = {node: [] for node in meta_dfs.keys()}

    # 4) Iterative removal + logging
    with open(log_txt, "w") as log:
        log.write("Entries to remove (and discovered children):\n")
        log.write("\n".join(pending) + "\n\n")

        while pending:
            curr = pending.pop(0)
            log.write(f"Removing: {curr}\n")
            for node, df in meta_dfs.items():
                node_id_col = f"{node}_id"
                # direct hits
                if node_id_col in df.columns:
                    hits = df[df[node_id_col] == curr].index.tolist()
                    if hits:
                        deleted[node].append(curr)
                        log.write(f"  - {curr} dropped from {node}.{node_id_col}\n")
                        # remove from meta_dfs and add to meta_dfs_delete
                        meta_dfs_delete[node] = pd.concat(
                            [meta_dfs_delete[node], df.loc[hits]]
                        )
                        df.drop(index=hits, inplace=True)

                # discover child entries
                link_cols = [c for c in df.columns if "." in c and c.endswith("_id")]
                for lc in link_cols:
                    matches = df[df[lc] == curr]
                    for _, row in matches.iterrows():
                        child = row[f"{node}_id"]
                        if child not in deleted[node] and child not in pending:
                            pending.append(child)
                            log.write(
                                f"    => discovered child {child} in {node}.{lc}\n"
                            )

        # summary
        log.write("\nSummary of deletions by sheet:\n")
        for node, items in deleted.items():
            log.write(f" {node}: {items}\n")

    # 5) Write out with openpyxl in-place with modified data of things kept (not removed). This ensures we keep any formatting, formulas, etc. that may be present in the original manifest.
    if file:
        logger.info(f"Writing cleaned manifest to {out_xlsx}")
        wb = load_workbook(manifest_path)
        for node, df in meta_dfs.items():
            if node in wb.sheetnames:
                ws = wb[node]
                ws.delete_rows(1, ws.max_row)
            else:
                ws = wb.create_sheet(title=node)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

        # ensure at least one sheet visible
        if not wb.sheetnames:
            wb.create_sheet(title="Sheet1")
        wb.save(out_xlsx)

        print(
            f"\n✅ Done. Log written to {log_txt}\n   Cleaned workbook: {out_xlsx}\n   Deleted entries workbook: {out_delete_xlsx}\n"
        )

    if directory:
        logger.info(
            f"Writing cleaned manifest to {out_folder} directory with individual TSVs for each node"
        )
        os.makedirs(out_folder, exist_ok=True)
        for node, df in meta_dfs.items():
            out_tsv_path = os.path.join(out_folder, f"{node}{base_out_tsv}")
            df.to_csv(out_tsv_path, sep="\t", index=False)

    # also write out the deleted entries for reference
    with pd.ExcelWriter(out_delete_xlsx, engine="openpyxl") as writer:
        for node, df in meta_dfs_delete.items():
            if not df.empty:
                df.to_excel(writer, sheet_name=node, index=False)

        print(
            f"\n✅ Done. Log written to {log_txt}\n   Cleaned TSVs: {out_folder}\n   Deleted entries workbook: {out_delete_xlsx}\n"
        )

    return out_xlsx, log_txt, out_delete_xlsx, out_folder


@flow(name="CCDI Manifest Entry Remover", flow_run_name="{runner}_" + f"{get_time()}")
def entry_remover(
    bucket: str,
    runner: str,
    file_path: str,
    directory_path: str,
    entry_removal_file_path: str,
) -> None:
    """
    Prefect flow to remove specified entries from a CCDI manifest, given either as a single Excel file or a directory of TSVs. The entries to remove are provided in a separate TSV file. The flow handles downloading the necessary files from cloud storage, performing the entry removal, and uploading the results back to cloud storage.

    Parameters:
    - bucket: The cloud storage bucket where the input files are located and where the output should be uploaded.
    - runner: A string identifier for the runner of the flow, used in naming output folders/files.
    - file_path: The path to the manifest file in the bucket (if using a single Excel file input).
    - directory_path: The path to the manifest directory in the bucket (if using a directory of TSVs input).
    - entry_removal_file_path: The path to the TSV file in the bucket that lists the entries to remove.
    """

    logger = get_run_logger()

    output_folder = runner + "/entry_remover_" + get_time()

    if file_path == "path/to/file":
        file_path = None
    if directory_path == "path/to/directory":
        directory_path = None

    if file_path:
        # download manifest
        logger.info(f"Downloading manifest from {file_path} in bucket {bucket}")
        file_dl(filename=file_path, bucket=bucket)
        file_path = os.path.basename(file_path)
        directory_path = None

    if directory_path:
        # download manifest directory
        logger.info(f"Downloading directory from {directory_path} in bucket {bucket}")
        folder_dl(bucket=bucket, remote_folder=directory_path)
        directory_path = os.path.basename(directory_path)
        file_path = None

    if entry_removal_file_path:
        # download tsv of entries to remove
        logger.info(
            f"Downloading entry removal file from {entry_removal_file_path} in bucket {bucket}"
        )
        file_dl(filename=entry_removal_file_path, bucket=bucket)
        entry_removal_file_path = os.path.basename(entry_removal_file_path)
    else:
        logger.warning("No entry removal file provided. Exiting.")
        return

    out_xlsx, log_txt, out_delete_xlsx, out_folder = main(
        file=file_path, directory=directory_path, entry=entry_removal_file_path
    )

    if out_xlsx:
        file_ul(
            newfile=out_xlsx, bucket=bucket, output_folder=output_folder, sub_folder=""
        )
    if log_txt:
        file_ul(
            newfile=log_txt, bucket=bucket, output_folder=output_folder, sub_folder=""
        )
    if out_delete_xlsx:
        file_ul(
            newfile=out_delete_xlsx,
            bucket=bucket,
            output_folder=output_folder,
            sub_folder="",
        )
    if out_folder:
        folder_ul(
            local_folder=out_folder,
            bucket=bucket,
            destination=output_folder,
            sub_folder="",
        )
